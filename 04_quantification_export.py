# =============================================================================
# Script 04 — Pore Network Quantification and Excel Export (v2)
# =============================================================================
# Author:      Andrea Tonelli (tnland001@myuct.ac.za)
# ORCID:       https://orcid.org/0000-0002-1601-4103
# Institution: University of Cape Town
# Repository:  https://github.com/andtoni/growth-tunnel-analysis
# Preprint:    https://dx.doi.org/10.2139/ssrn.6664677
#
# Description:
#   Loads saved SNOW2 outputs and applies identical thresholding to extract
#   topology- and density-based metrics that meaningfully characterise the
#   thresholded pore network independent of the threshold value itself.
#
#   Mean pore/throat diameter are deliberately excluded as primary metrics
#   because they are trivially bounded by the threshold. Instead, this script
#   reports network density, topological connectivity (Euler/coordination),
#   and geometric ratios that are biologically interpretable and threshold-
#   sensitivity demonstrable.
#
#   Rebuilding from pkl is FAST (seconds) — SNOW2 is never re-run.
#
# Primary metrics reported:
#   Porosity (%)                — volume-based, threshold-independent
#   Pore density (pores/mm³)   — normalised, analogous to vascular density
#   Throat density (throats/mm³)
#   Connectivity density (mm⁻³) — Euler-based, analogous to bone literature
#   Coordination number         — KEY topology metric
#   Dead-end pore fraction (%)  — non-permissive pores
#   Well-connected pore % (≥3)  — angio-permissive pores
#   Mean throat length ± SD     — migration distance between pores
#   Throat aspect ratio         — length/diameter, migration difficulty index
#   Constriction ratio          — pore/throat diameter, constraint metric
#
# Usage:
#   python scripts/04_quantification_export.py
#
# Requires:
#   openpyxl — pip install openpyxl
#   networkx — pip install networkx  (for path length — optional)
#
# Excel workbook structure:
#   Sheet 1  README                     — how to use in GraphPad Prism
#   Sheet 2  Primary Metrics Summary    — key metrics, one row per run
#   Sheet 3  Threshold Sensitivity      — metrics vs threshold (XY plots)
#   Sheet 4  Coordination Numbers       — Prism column table
#   Sheet 5  Throat Lengths             — Prism column table
#   Sheet 6  Throat Aspect Ratios       — Prism column table
#   Sheet 7  Constriction Ratios        — Prism column table
#   Sheet 8  Full Data Archive          — all raw arrays (incl. diameters)
#   Sheet 9  Metric Descriptions        — full descriptions, units, biology
# =============================================================================

import matplotlib
matplotlib.use('Agg')

import porespy as ps
import openpnm as op
import numpy as np
import pandas as pd
import pickle
import os
import sys
import glob
from datetime import datetime
from scipy.sparse import csr_matrix
from scipy.sparse.csgraph import connected_components
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# USER SETTINGS
# =============================================================================

# Must match base_data_dir in run_snow2_only.py and run_network_analysis_v2.py
base_data_dir = r"C:\Users\andto\OneDrive\Desktop\University\PhD\DATA\Transmural Space Characterisation\3D Analysis Paper\codeoutput"

# Sample names — must match sample_name values in the original scripts
samples = [
    "SR-Pel16",
    "SR-Pel18",
    "SR-Pel20",
]

# Threshold combinations (pore_threshold, throat_threshold) in µm
threshold_combinations = [
    (5,    5),
    (7.5,  7.5),
    (10,   10),
    (12.5, 12.5),
    (15,   15),
]

# Voxel size in µm — from your scan acquisition parameters
# Used to calculate scaffold volume for density metrics
voxel_size_um = 0.54

# Output Excel filename — saved to base_data_dir
excel_filename = "Pore_Network_Quantification.xlsx"

# Compute mean shortest path length (requires networkx, slow for large nets)
# Set False if networkx is not installed or networks are very large
compute_path_length = False

# =============================================================================
# DO NOT EDIT BELOW THIS LINE
# =============================================================================

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
    sys.stdout.flush()

def build_run_label(pore_t, throat_t):
    return f"pore{pore_t}um_throat{throat_t}um"

def extract_largest_cluster(pn):
    """Replicates run_network_analysis_v2.py cluster extraction exactly."""
    n = pn.num_pores()
    if n == 0 or len(pn['throat.conns']) == 0:
        return
    conns = pn['throat.conns']
    adj   = csr_matrix(
        (np.ones(len(conns)), (conns[:, 0], conns[:, 1])),
        shape=(n, n)
    )
    n_comp, labels = connected_components(adj, directed=False, connection='weak')
    if n_comp == 1:
        return
    sizes = np.bincount(labels)
    op.topotools.trim(
        network=pn,
        pores=np.where(labels != int(np.argmax(sizes)))[0]
    )

def rebuild_network(pkl_path, im_pores_path, pore_t, throat_t):
    """
    Loads pkl, rebuilds OpenPNM network, applies identical thresholding
    to run_network_analysis_v2.py. Returns (pn, porosity, vol_mm3).
    """
    with open(pkl_path, "rb") as f:
        snow_output = pickle.load(f)
    im = np.load(im_pores_path)

    # Scaffold volume in mm³
    shape      = im.shape
    vol_um3    = shape[0] * shape[1] * shape[2] * (voxel_size_um ** 3)
    vol_mm3    = vol_um3 / 1e9
    porosity   = ps.metrics.porosity(im) * 100

    pn = op.io.network_from_porespy(snow_output.network)
    geo = op.models.collections.geometry.spheres_and_cylinders
    pn.add_model_collection(geo, domain='all')
    pn.regenerate_models()

    # Threshold — mirrors run_network_analysis_v2.py exactly
    op.topotools.trim(
        pn, throats=pn["throat.inscribed_diameter"] < throat_t
    )
    op.topotools.trim(
        pn, pores=pn["pore.inscribed_diameter"] < pore_t
    )
    extract_largest_cluster(pn)

    for _ in range(10):
        h     = op.utils.check_network_health(pn)
        trim  = np.union1d(h['isolated_pores'], h['disconnected_pores'])
        if len(trim) == 0:
            break
        op.topotools.trim(network=pn, pores=trim)

    extract_largest_cluster(pn)
    return pn, porosity, vol_mm3

def compute_constriction_ratios(pn):
    """
    Per-pore ratio of pore inscribed diameter to mean diameter of its
    connecting throats. Captures the tightness of the pore-throat junction.
    """
    conns        = pn['throat.conns']
    pore_diams   = pn['pore.inscribed_diameter']
    throat_diams = pn['throat.inscribed_diameter']
    ratios = []
    for p in range(pn.num_pores()):
        t_idx = np.where((conns[:, 0] == p) | (conns[:, 1] == p))[0]
        if len(t_idx) > 0:
            ratios.append(pore_diams[p] / np.mean(throat_diams[t_idx]))
    return np.array(ratios)

def compute_mean_path_length(pn):
    """
    Mean shortest path length in hops across the network.
    Requires networkx. Returns None if networkx unavailable or network too large.
    """
    try:
        import networkx as nx
        G = nx.Graph()
        G.add_nodes_from(range(pn.num_pores()))
        G.add_edges_from(pn['throat.conns'].tolist())
        if nx.is_connected(G):
            # Sample if large to keep runtime manageable
            if pn.num_pores() > 2000:
                nodes = np.random.choice(pn.num_pores(), 200, replace=False)
                lengths = []
                for n in nodes:
                    sp = nx.single_source_shortest_path_length(G, n)
                    lengths.extend(sp.values())
                return round(float(np.mean(lengths)), 3)
            else:
                return round(nx.average_shortest_path_length(G), 3)
        else:
            return None
    except ImportError:
        return None

def compute_metrics(pn, porosity, vol_mm3, pore_t, throat_t):
    """
    Computes all primary quantitative metrics for one run.
    Returns (summary dict, arrays dict).
    """
    n_pores   = pn.num_pores()
    n_throats = pn.num_throats()

    # ── Core arrays ──────────────────────────────────────────────────────────
    coord_nums       = pn['pore.coordination_number']
    throat_diams     = pn['throat.inscribed_diameter']
    throat_lens      = pn['throat.length']
    pore_diams       = pn['pore.inscribed_diameter']
    pore_vols        = pn['pore.volume']
    throat_ar        = throat_lens / throat_diams           # aspect ratio
    constriction_r   = compute_constriction_ratios(pn)

    # ── Euler characteristic and connectivity density ─────────────────────────
    # For a connected pore network graph:
    #   χ = N_pores - N_throats  (Euler characteristic, ignoring loops)
    #   β = 1 - χ = N_throats - N_pores + 1  (connectivity = redundant loops)
    #   Connectivity density = β / vol_mm³
    euler_char          = n_pores - n_throats
    connectivity        = n_throats - n_pores + 1
    connectivity_density = connectivity / vol_mm3 if vol_mm3 > 0 else 0

    # ── Path length (optional) ────────────────────────────────────────────────
    path_len = compute_mean_path_length(pn) if compute_path_length else None

    summary = {
        # ── Identifiers ──────────────────────────────────────────────────────
        'sample':                    None,      # filled by caller
        'threshold_um':              pore_t,
        'run_label':                 None,

        # ── Volume ───────────────────────────────────────────────────────────
        'scaffold_volume_mm3':       round(vol_mm3, 6),
        'porosity_pct':              round(porosity, 3),

        # ── Density (normalised per mm³) ──────────────────────────────────────
        'pore_density_per_mm3':      round(n_pores  / vol_mm3, 2) if vol_mm3 > 0 else 0,
        'throat_density_per_mm3':    round(n_throats / vol_mm3, 2) if vol_mm3 > 0 else 0,

        # ── Topology ──────────────────────────────────────────────────────────
        'pore_count':                int(n_pores),
        'throat_count':              int(n_throats),
        'throat_to_pore_ratio':      round(n_throats / n_pores, 3) if n_pores > 0 else 0,
        'euler_characteristic':      int(euler_char),
        'connectivity':              int(connectivity),
        'connectivity_density_mm3':  round(connectivity_density, 2),

        # ── Coordination number ───────────────────────────────────────────────
        'coord_num_n':               int(n_pores),
        'coord_num_mean':            round(float(np.mean(coord_nums)), 3),
        'coord_num_sd':              round(float(np.std(coord_nums)),  3),
        'coord_num_median':          round(float(np.median(coord_nums)), 1),
        'dead_end_pores_pct':        round(float((coord_nums == 1).sum() / n_pores * 100), 2),
        'well_connected_pores_pct':  round(float((coord_nums >= 3).sum() / n_pores * 100), 2),

        # ── Throat geometry ───────────────────────────────────────────────────
        'throat_length_n':           int(n_throats),
        'throat_length_mean_um':     round(float(np.mean(throat_lens)), 3),
        'throat_length_sd_um':       round(float(np.std(throat_lens)),  3),
        'throat_length_median_um':   round(float(np.median(throat_lens)), 3),
        'throat_aspect_ratio_n':     int(n_throats),
        'throat_aspect_ratio_mean':  round(float(np.mean(throat_ar)),  3),
        'throat_aspect_ratio_sd':    round(float(np.std(throat_ar)),   3),

        # ── Constriction ratio ────────────────────────────────────────────────
        'constriction_ratio_n':      int(len(constriction_r)),
        'constriction_ratio_mean':   round(float(np.mean(constriction_r)),  3) if len(constriction_r) > 0 else None,
        'constriction_ratio_sd':     round(float(np.std(constriction_r)),   3) if len(constriction_r) > 0 else None,
        'constriction_ratio_median': round(float(np.median(constriction_r)), 3) if len(constriction_r) > 0 else None,

        # ── Path length (optional) ────────────────────────────────────────────
        'mean_shortest_path_hops':   path_len,

        # ── Archived: diameter data (not primary metrics) ─────────────────────
        'pore_diam_mean_um':         round(float(np.mean(pore_diams)), 3),
        'pore_diam_sd_um':           round(float(np.std(pore_diams)),  3),
        'throat_diam_mean_um':       round(float(np.mean(throat_diams)), 3),
        'throat_diam_sd_um':         round(float(np.std(throat_diams)),  3),
        'pore_volume_mean_um3':      round(float(np.mean(pore_vols)), 3),
        'pore_volume_total_um3':     round(float(np.sum(pore_vols)),  3),
    }

    arrays = {
        'coord_nums':      coord_nums.astype(float),
        'throat_lens':     throat_lens,
        'throat_ar':       throat_ar,
        'constriction_r':  constriction_r,
        'pore_diams':      pore_diams,
        'throat_diams':    throat_diams,
        'pore_vols':       pore_vols,
    }

    return summary, arrays

# =============================================================================
# COLLECT ALL DATA
# =============================================================================

log("=" * 65)
log("Script 04 — Pore Network Quantification Export (v2)")
log("Tonelli A. — University of Cape Town — 2025")
log("=" * 65)
log(f"Base directory:  {base_data_dir}")
log(f"Samples:         {samples}")
log(f"Thresholds:      {threshold_combinations}")
log(f"Voxel size:      {voxel_size_um} µm")
log("=" * 65)

all_summaries = []
all_arrays    = {}
skipped       = []
errors        = []

for sample in samples:
    sample_dir    = os.path.join(base_data_dir, sample)
    pkl_path      = os.path.join(sample_dir, "snow2_output.pkl")
    im_pores_path = os.path.join(sample_dir, "im_pores.npy")

    for check, name in [(pkl_path, "snow2_output.pkl"),
                         (im_pores_path, "im_pores.npy")]:
        if not os.path.exists(check):
            log(f"SKIPPING {sample} — {name} not found")
            skipped.append((sample, "all", f"{name} missing"))
            break
    else:
        for pore_t, throat_t in threshold_combinations:
            run_label = build_run_label(pore_t, throat_t)
            log(f"Processing: {sample} / {run_label}")

            try:
                pn, porosity, vol_mm3 = rebuild_network(
                    pkl_path, im_pores_path, pore_t, throat_t
                )
                summary, arrays = compute_metrics(
                    pn, porosity, vol_mm3, pore_t, throat_t
                )
                summary['sample']    = sample
                summary['run_label'] = run_label
                all_summaries.append(summary)

                group_label = f"{sample}_{pore_t}µm"
                all_arrays[group_label] = arrays

                log(f"  Vol: {vol_mm3:.4f} mm³ | "
                    f"Pore density: {summary['pore_density_per_mm3']:.0f}/mm³ | "
                    f"Coord: {summary['coord_num_mean']:.2f} | "
                    f"Well-conn: {summary['well_connected_pores_pct']:.1f}%")

            except Exception as e:
                log(f"  ERROR: {e}")
                import traceback; traceback.print_exc()
                errors.append((sample, run_label, str(e)))

if not all_summaries:
    log("ERROR: No data collected. Check base_data_dir and sample names.")
    sys.exit(1)

df = pd.DataFrame(all_summaries)
log(f"\nCollected {len(all_summaries)} runs")

# =============================================================================
# EXCEL STYLE CONSTANTS
# =============================================================================

FONT       = "Arial"
C_HDR      = "1F4E79"   # dark navy
C_SUB      = "2E75B6"   # mid blue
C_ALT      = "D6E4F0"   # light blue rows
C_DENSITY  = "E2EFDA"   # green — density / volume metrics
C_TOPO     = "FCE4D6"   # orange — topology metrics
C_COORD    = "FFF2CC"   # yellow — coordination
C_THROAT   = "EBF3FB"   # pale blue — throat geometry
C_CONSTR   = "F4EBFF"   # lavender — constriction
C_ARCHIVE  = "F2F2F2"   # grey — archived / secondary metrics
C_WHITE    = "FFFFFF"

def hc(ws, row, col, val, bg=C_HDR, fg="FFFFFF",
        bold=True, sz=10, wrap=True, left=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name=FONT, bold=bold, color=fg, size=sz)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="center", wrap_text=wrap
    )
    return c

def dc(ws, row, col, val, bg=C_WHITE, fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name=FONT, size=9, bold=bold)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c

def freeze(ws, row, col):
    ws.freeze_panes = f"{get_column_letter(col)}{row}"

def cw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# =============================================================================
# SHEET 1 — README
# =============================================================================

wb  = Workbook()
wb.remove(wb.active)
ws1 = wb.create_sheet("README")
ws1.sheet_view.showGridLines = False

hc(ws1, 1, 1,
   "Pore Network Quantification — GraphPad Prism Import Guide",
   bg=C_HDR, fg="FFFFFF", sz=13, left=True)
ws1.merge_cells("A1:D1")
ws1.row_dimensions[1].height = 32

info = [
    ("", ""),
    ("Author",      "Andrea Tonelli — University of Cape Town"),
    ("ORCID",       "https://orcid.org/0000-0002-1601-4103"),
    ("Repository",  "https://github.com/andtoni/growth-tunnel-analysis"),
    ("Preprint",    "https://dx.doi.org/10.2139/ssrn.6664677"),
    ("Zenodo DOI",  "https://doi.org/10.5281/zenodo.20308262"),
    ("Generated",   datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("", ""),
    ("METRIC RATIONALE", ""),
    ("Why not mean diameter?",
     "Mean pore/throat diameter after thresholding is trivially bounded by the "
     "threshold value — the minimum always equals the threshold. Reporting it as "
     "a primary metric is circular. Diameter data is retained in Sheet 8 (Full "
     "Data Archive) for completeness but is not a primary result."),
    ("Primary metrics used",
     "Density (pores/mm³, throats/mm³), topological connectivity (Euler "
     "characteristic, coordination number), migration geometry (throat length, "
     "aspect ratio, constriction ratio). These are threshold-sensitivity "
     "demonstrable and biologically interpretable."),
    ("", ""),
    ("SHEET GUIDE", ""),
    ("Sheet 2 — Primary Metrics Summary",
     "One row per sample × threshold. All primary metrics. "
     "Import into Prism as Grouped table for bar charts or heatmaps."),
    ("Sheet 3 — Threshold Sensitivity",
     "Primary metrics vs threshold, one column per sample. "
     "Import as XY Table (X = threshold, Y = metric) for sensitivity curves. "
     "A plateau region justifies your chosen threshold."),
    ("Sheet 4 — Coordination Numbers",
     "Column table: one column per group (Sample_Threshold). "
     "Import into Prism → Column table → box plots or frequency distribution. "
     "KEY metric for angio-permissive characterisation."),
    ("Sheet 5 — Throat Lengths",
     "Column table: one column per group. "
     "Import into Prism → box plots. Migration distance between pores."),
    ("Sheet 6 — Throat Aspect Ratios",
     "Column table: one column per group. Throat length / throat diameter. "
     "Migration difficulty index — high values = long narrow passages."),
    ("Sheet 7 — Constriction Ratios",
     "Column table: one column per group. Pore diameter / mean throat diameter. "
     "High values = wide chambers with narrow exits (constrains cell transit)."),
    ("Sheet 8 — Full Data Archive",
     "Pore diameters, throat diameters, pore volumes per group. "
     "Retained for completeness — treat as supplementary, not primary results."),
    ("Sheet 9 — Metric Descriptions",
     "Full description, unit, and biological significance of every metric."),
    ("", ""),
    ("HOW TO IMPORT COLUMN TABLE SHEETS INTO PRISM", ""),
    ("Step 1", "Open GraphPad Prism → New → Column Table"),
    ("Step 2", "File → Import → From File → select this Excel file → choose sheet"),
    ("Step 3",
     "Each column becomes a group. Use for box plots, violin plots, "
     "or frequency distribution analysis."),
    ("Step 4",
     "For threshold sensitivity (Sheet 3): New → XY Table → "
     "X = Threshold (µm), Y = metric. One dataset per sample."),
]

r = 2
for label, val in info:
    bold = label in ("METRIC RATIONALE", "SHEET GUIDE",
                     "HOW TO IMPORT COLUMN TABLE SHEETS INTO PRISM")
    bg   = C_SUB if bold else None
    fg_c = "FFFFFF" if bold else "000000"
    for col, v in ((1, label), (2, val)):
        c = ws1.cell(row=r, column=col, value=v)
        c.font      = Font(name=FONT, bold=bold, color=fg_c, size=9)
        c.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
    ws1.row_dimensions[r].height = 40 if len(str(val)) > 100 else 20
    r += 1

cw(ws1, 1, 30); cw(ws1, 2, 90)

# =============================================================================
# SHEET 2 — PRIMARY METRICS SUMMARY
# =============================================================================

ws2 = wb.create_sheet("Primary Metrics Summary")
ws2.sheet_view.showGridLines = False

# Column definitions: (header, df_key, width, bg_colour, number_format)
col_defs = [
    # Identifiers
    ("Sample",                      "sample",                   18, C_HDR,    None),
    ("Threshold (µm)",              "threshold_um",             13, C_HDR,    "0.0"),

    # Volume & porosity
    ("Scaffold Vol (mm³)",          "scaffold_volume_mm3",      15, C_DENSITY, "0.000000"),
    ("Porosity (%)",                "porosity_pct",             12, C_DENSITY, "0.000"),

    # Density — PRIMARY
    ("Pore Density (/mm³)",         "pore_density_per_mm3",     15, C_DENSITY, "#,##0.0"),
    ("Throat Density (/mm³)",       "throat_density_per_mm3",   15, C_DENSITY, "#,##0.0"),

    # Topology — PRIMARY
    ("Pore Count",                  "pore_count",               11, C_TOPO,   "#,##0"),
    ("Throat Count",                "throat_count",             12, C_TOPO,   "#,##0"),
    ("Throat:Pore Ratio",           "throat_to_pore_ratio",     14, C_TOPO,   "0.000"),
    ("Euler Characteristic χ",      "euler_characteristic",     16, C_TOPO,   "0"),
    ("Connectivity β",              "connectivity",             13, C_TOPO,   "0"),
    ("Connectivity Density (/mm³)", "connectivity_density_mm3", 20, C_TOPO,   "#,##0.0"),

    # Coordination — PRIMARY
    ("Coord Num n",                 "coord_num_n",              10, C_COORD,  "#,##0"),
    ("Coord Num Mean",              "coord_num_mean",           14, C_COORD,  "0.000"),
    ("Coord Num SD",                "coord_num_sd",             12, C_COORD,  "0.000"),
    ("Coord Num Median",            "coord_num_median",         14, C_COORD,  "0.0"),
    ("Dead-End Pores (%)",          "dead_end_pores_pct",       14, C_COORD,  "0.00"),
    ("Well-Connected ≥3 (%)",       "well_connected_pores_pct", 16, C_COORD,  "0.00"),

    # Throat geometry — PRIMARY
    ("Throat Length n",             "throat_length_n",          12, C_THROAT, "#,##0"),
    ("Throat Length Mean (µm)",     "throat_length_mean_um",    18, C_THROAT, "0.000"),
    ("Throat Length SD (µm)",       "throat_length_sd_um",      16, C_THROAT, "0.000"),
    ("Throat Aspect Ratio n",       "throat_aspect_ratio_n",    15, C_THROAT, "#,##0"),
    ("Throat Aspect Ratio Mean",    "throat_aspect_ratio_mean", 19, C_THROAT, "0.000"),
    ("Throat Aspect Ratio SD",      "throat_aspect_ratio_sd",   17, C_THROAT, "0.000"),

    # Constriction — PRIMARY
    ("Constriction Ratio n",        "constriction_ratio_n",     15, C_CONSTR, "#,##0"),
    ("Constriction Ratio Mean",     "constriction_ratio_mean",  18, C_CONSTR, "0.000"),
    ("Constriction Ratio SD",       "constriction_ratio_sd",    16, C_CONSTR, "0.000"),
    ("Constriction Ratio Median",   "constriction_ratio_median",18, C_CONSTR, "0.000"),

    # Path length (optional)
    ("Mean Shortest Path (hops)",   "mean_shortest_path_hops",  20, C_TOPO,  "0.000"),

    # Archived diameter data
    ("Pore Diam Mean (µm)*",        "pore_diam_mean_um",        16, C_ARCHIVE,"0.000"),
    ("Pore Diam SD (µm)*",          "pore_diam_sd_um",          14, C_ARCHIVE,"0.000"),
    ("Throat Diam Mean (µm)*",      "throat_diam_mean_um",      17, C_ARCHIVE,"0.000"),
    ("Throat Diam SD (µm)*",        "throat_diam_sd_um",        15, C_ARCHIVE,"0.000"),
]

ws2.row_dimensions[1].height = 50
for ci, (hdr, key, width, clr, fmt) in enumerate(col_defs, 1):
    fg = "FFFFFF" if clr in (C_HDR,) else "000000"
    hc(ws2, 1, ci, hdr, bg=clr, fg=fg, sz=9, wrap=True)
    cw(ws2, ci, width)

# Add footnote marker row for archived columns
note = ws2.cell(row=2, column=1,
    value="* Diameter metrics shown for reference only — not primary results "
          "(mean is trivially bounded by threshold value)")
note.font      = Font(name=FONT, italic=True, size=8, color="595959")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws2.merge_cells(f"A2:{get_column_letter(len(col_defs))}2")
ws2.row_dimensions[2].height = 20

for ri, (_, row) in enumerate(df.iterrows()):
    er  = ri + 3
    bg  = C_ALT if ri % 2 == 0 else C_WHITE
    for ci, (hdr, key, width, clr, fmt) in enumerate(col_defs, 1):
        val = row.get(key, None)
        c   = dc(ws2, er, ci, val, bg=bg, fmt=fmt)

freeze(ws2, 3, 3)
ws2.auto_filter.ref = f"A1:{get_column_letter(len(col_defs))}1"

# =============================================================================
# SHEET 3 — THRESHOLD SENSITIVITY
# =============================================================================

ws3 = wb.create_sheet("Threshold Sensitivity")
ws3.sheet_view.showGridLines = False

hc(ws3, 1, 1,
   "Threshold Sensitivity Analysis — Import as XY Table in Prism "
   "(X = Threshold µm, Y = metric value) — plateau region justifies threshold choice",
   bg=C_SUB, fg="FFFFFF", sz=10, left=True)
ws3.merge_cells("A1:Z1")
ws3.row_dimensions[1].height = 28

sens_metrics = [
    ("Pore Density (/mm³)",          "pore_density_per_mm3",     C_DENSITY, "#,##0.0"),
    ("Throat Density (/mm³)",        "throat_density_per_mm3",   C_DENSITY, "#,##0.0"),
    ("Connectivity Density (/mm³)",  "connectivity_density_mm3", C_TOPO,   "#,##0.0"),
    ("Coordination Num Mean",        "coord_num_mean",           C_COORD,  "0.000"),
    ("Dead-End Pores (%)",           "dead_end_pores_pct",       C_COORD,  "0.00"),
    ("Well-Connected ≥3 (%)",        "well_connected_pores_pct", C_COORD,  "0.00"),
    ("Throat Length Mean (µm)",      "throat_length_mean_um",    C_THROAT, "0.000"),
    ("Throat Aspect Ratio Mean",     "throat_aspect_ratio_mean", C_THROAT, "0.000"),
    ("Constriction Ratio Mean",      "constriction_ratio_mean",  C_CONSTR, "0.000"),
    ("Porosity (%)",                 "porosity_pct",             C_DENSITY,"0.000"),
]

thresholds = sorted(df['threshold_um'].unique())
start_col  = 1

for metric_label, metric_key, clr, fmt in sens_metrics:
    n_cols = len(samples) + 1
    end_col = start_col + n_cols - 1

    hc(ws3, 2, start_col, metric_label, bg=clr,
       fg="000000", sz=9, wrap=True)
    if n_cols > 1:
        ws3.merge_cells(
            f"{get_column_letter(start_col)}2:"
            f"{get_column_letter(end_col)}2"
        )

    hc(ws3, 3, start_col, "Threshold (µm)", bg=C_HDR, fg="FFFFFF", sz=9)
    cw(ws3, start_col, 14)
    for si, samp in enumerate(samples):
        hc(ws3, 3, start_col + si + 1, samp, bg=C_HDR, fg="FFFFFF", sz=9)
        cw(ws3, start_col + si + 1, 16)

    for ri, thresh in enumerate(thresholds):
        er = 4 + ri
        bg = C_ALT if ri % 2 == 0 else C_WHITE
        dc(ws3, er, start_col, thresh, bg=bg, fmt="0.0")
        for si, samp in enumerate(samples):
            mask = ((df['sample'] == samp) &
                    (df['threshold_um'] == thresh))
            vals = df.loc[mask, metric_key].values
            v = float(vals[0]) if len(vals) > 0 and vals[0] is not None else None
            dc(ws3, er, start_col + si + 1, v, bg=bg, fmt=fmt)

    start_col += n_cols + 2

ws3.row_dimensions[2].height = 30
ws3.row_dimensions[3].height = 35

# =============================================================================
# HELPER — PRISM COLUMN TABLE SHEET
# =============================================================================

def prism_sheet(wb, name, array_key, unit, title, description, clr):
    ws  = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    sorted_groups = sorted(all_arrays.keys())
    n_grp = len(sorted_groups)

    hc(ws, 1, 1, title, bg=C_SUB, fg="FFFFFF", sz=10, left=True)
    if n_grp > 1:
        ws.merge_cells(f"A1:{get_column_letter(n_grp)}1")
    ws.row_dimensions[1].height = 25

    c = ws.cell(row=2, column=1, value=description)
    c.font      = Font(name=FONT, italic=True, size=8, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True)
    if n_grp > 1:
        ws.merge_cells(f"A2:{get_column_letter(n_grp)}2")
    ws.row_dimensions[2].height = 35

    for ci, grp in enumerate(sorted_groups, 1):
        hc(ws, 3, ci, grp, bg=clr, fg="000000", sz=9)
        cw(ws, ci, 20)

    for ci in range(1, n_grp + 1):
        c = ws.cell(row=4, column=ci, value=unit)
        c.font      = Font(name=FONT, italic=True, size=8, color="595959")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = PatternFill("solid", fgColor="F2F2F2")
    ws.row_dimensions[4].height = 14

    max_len = max(len(all_arrays[g][array_key]) for g in sorted_groups)
    for ri in range(max_len):
        er = 5 + ri
        bg = C_ALT if ri % 2 == 0 else C_WHITE
        for ci, grp in enumerate(sorted_groups, 1):
            arr = all_arrays[grp][array_key]
            val = float(arr[ri]) if ri < len(arr) else None
            c   = ws.cell(row=er, column=ci, value=val)
            c.font      = Font(name=FONT, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill      = PatternFill("solid", fgColor=bg)
            if val is not None:
                c.number_format = "0.000"

    freeze(ws, 5, 1)

# =============================================================================
# SHEETS 4–7 — PRISM COLUMN TABLES (PRIMARY METRICS)
# =============================================================================

log("Writing Prism column table sheets...")

prism_sheet(wb, "Coordination Numbers", "coord_nums",
    unit        = "connections/pore",
    title       = "Pore Coordination Number  |  KEY METRIC for angio-permissive characterisation",
    description = ("One column per group (Sample_Threshold). Import into Prism as Column Table → "
                   "box plots or frequency distribution. "
                   "Coord = 1: dead-end (no exit route). "
                   "Coord ≥ 3: branching network (supports anastomosis)."),
    clr         = C_COORD)

prism_sheet(wb, "Throat Lengths", "throat_lens",
    unit        = "µm",
    title       = "Throat Length (µm)  |  Migration distance between pores",
    description = ("One column per group. Import into Prism → box plots. "
                   "Throat length is the centre-to-centre distance between connected pores. "
                   "Shorter throats permit faster cell migration."),
    clr         = C_THROAT)

prism_sheet(wb, "Throat Aspect Ratios", "throat_ar",
    unit        = "length / diameter",
    title       = "Throat Aspect Ratio (Length/Diameter)  |  Migration difficulty index",
    description = ("One column per group. Import into Prism → box plots. "
                   "Higher values = longer and narrower throats = more difficult cell transit. "
                   "Ratio > 5 suggests a constrained migration pathway."),
    clr         = C_THROAT)

prism_sheet(wb, "Constriction Ratios", "constriction_r",
    unit        = "pore diam / mean throat diam",
    title       = "Constriction Ratio (Pore/Throat Diameter)  |  Pore-throat size contrast",
    description = ("One column per group. Import into Prism → box plots. "
                   "Per pore: ratio of pore inscribed diameter to mean diameter of its connecting throats. "
                   "High values = wide chambers with narrow exits → constrains cell entry/exit."),
    clr         = C_CONSTR)

# =============================================================================
# SHEET 8 — FULL DATA ARCHIVE (diameter data)
# =============================================================================

prism_sheet(wb, "Pore Diameters Archive", "pore_diams",
    unit        = "µm",
    title       = "Pore Inscribed Diameter — ARCHIVED (not primary metric — see README)",
    description = ("Mean is trivially bounded by the threshold value. "
                   "Retained for completeness and reproducibility. "
                   "Do not use as a primary comparative result."),
    clr         = C_ARCHIVE)

prism_sheet(wb, "Throat Diameters Archive", "throat_diams",
    unit        = "µm",
    title       = "Throat Inscribed Diameter — ARCHIVED (not primary metric — see README)",
    description = ("Mean is trivially bounded by the threshold value. "
                   "Retained for completeness. Do not use as a primary result."),
    clr         = C_ARCHIVE)

prism_sheet(wb, "Pore Volumes Archive", "pore_vols",
    unit        = "µm³",
    title       = "Pore Volume — ARCHIVED",
    description = ("Retained for completeness. "
                   "Pore density (pores/mm³) is preferred for inter-sample comparison."),
    clr         = C_ARCHIVE)

# =============================================================================
# SHEET 9 — METRIC DESCRIPTIONS
# =============================================================================

ws9 = wb.create_sheet("Metric Descriptions")
ws9.sheet_view.showGridLines = False

hc(ws9, 1, 1, "Metric Descriptions — Primary Pore Network Metrics",
   bg=C_HDR, fg="FFFFFF", sz=12, left=True)
ws9.merge_cells("A1:E1")
ws9.row_dimensions[1].height = 28

for ci, hdr in enumerate(
        ["Metric", "Unit", "Description",
         "Biological Significance", "Reviewer Relevance"], 1):
    hc(ws9, 2, ci, hdr, bg=C_SUB, fg="FFFFFF", sz=10)

cw(ws9, 1, 26); cw(ws9, 2, 18)
cw(ws9, 3, 44); cw(ws9, 4, 50); cw(ws9, 5, 40)

metrics_table = [
    # (Metric, Unit, Description, Biological significance, Reviewer relevance)
    ("Porosity",
     "%",
     "Volume fraction of void (pore) space in the scaffold. "
     "Computed from the binary image — independent of thresholding.",
     "Higher porosity supports greater cell infiltration and vascularisation. "
     "A fundamental scaffold characterisation parameter.",
     "Validates binarisation and provides a threshold-independent baseline."),

    ("Pore Density",
     "pores/mm³",
     "Number of retained pores per cubic millimetre of scaffold. "
     "Normalises pore count to scaffold volume.",
     "Analogous to vascular density in histology — directly comparable "
     "across samples of different REV sizes. "
     "Higher density = more discrete growth spaces per unit volume.",
     "Threshold-sensitive: shows how network fragments with increasing threshold. "
     "Plateau region justifies threshold choice."),

    ("Throat Density",
     "throats/mm³",
     "Number of retained throats (pore connections) per cubic millimetre.",
     "Reflects interconnection density — how many migration passages "
     "are available per unit volume of scaffold.",
     "Complements pore density — together characterise network richness."),

    ("Euler Characteristic χ",
     "dimensionless",
     "Graph-theoretic topology measure: χ = N_pores − N_throats. "
     "More negative = more loops and redundant pathways in the network.",
     "Captures network complexity beyond simple counts. "
     "Used in trabecular bone and vascular network analysis.",
     "Threshold-independent topology metric comparable to bone literature."),

    ("Connectivity β",
     "dimensionless",
     "Number of independent loops in the network: β = N_throats − N_pores + 1. "
     "A tree (no loops) has β = 0. Each additional loop adds 1.",
     "Higher connectivity = more redundant migration pathways = "
     "greater resilience for cell migration if individual throats are blocked.",
     "Single number capturing topological robustness of the network."),

    ("Connectivity Density",
     "mm⁻³",
     "Connectivity (β) normalised to scaffold volume. "
     "Standard metric in trabecular bone microstructure analysis (Tb.N equivalent).",
     "Higher connectivity density = more branching pathways per unit volume = "
     "greater angio-permissive topology.",
     "Most powerful topology metric for reviewer response — "
     "directly comparable across scaffold groups and published bone/vascular data."),

    ("Coordination Number",
     "connections/pore",
     "Number of throat connections per pore — the node degree in the "
     "pore network graph.",
     "Coord = 1: dead-end pore (cell enters but cannot exit). "
     "Coord = 2: linear chain. "
     "Coord ≥ 3: branching node — supports anastomosis and network formation. "
     "Mean coordination number is the single best characterisation of "
     "angio-permissive network topology.",
     "KEY METRIC. Independent of threshold once above minimum — "
     "directly quantifies biological migration capacity."),

    ("Dead-End Pore Fraction",
     "%",
     "Percentage of pores with coordination number = 1.",
     "Dead-end pores trap cells and impair vascular network formation. "
     "High dead-end fraction indicates a poorly permissive scaffold.",
     "Identifies topologically non-permissive scaffold regions. "
     "Should decrease with increasing scaffold interconnectivity."),

    ("Well-Connected Pore Fraction",
     "%",
     "Percentage of pores with coordination number ≥ 3.",
     "Well-connected pores support branching migration and anastomosis — "
     "the hallmark of angio-permissive architecture.",
     "Direct quantification of angio-permissive topology. "
     "Primary metric for the paper's central argument."),

    ("Throat Length",
     "µm",
     "Centre-to-centre distance between pores connected by each throat. "
     "Reported as mean ± SD.",
     "Represents the migration distance a cell must traverse between pore spaces. "
     "Shorter throats facilitate faster and less constrained migration.",
     "Biologically interpretable — compare against cell migration speed data."),

    ("Throat Aspect Ratio",
     "length / diameter",
     "Ratio of throat length to throat inscribed diameter per throat.",
     "A high aspect ratio indicates a long narrow passage — "
     "physically demanding for cell migration. "
     "Ratio > 5 suggests a highly constrained migration corridor.",
     "Captures migration difficulty independent of absolute diameter. "
     "Useful for comparing scaffolds with different absolute pore sizes."),

    ("Constriction Ratio",
     "pore diam / throat diam",
     "Per-pore ratio of pore inscribed diameter to the mean diameter of "
     "its connecting throats.",
     "High ratio = wide pores with narrow constrictions — "
     "cells can enter and reside in pores but struggle to transit between them. "
     "Ratio close to 1 = smooth continuous pore space with few constrictions.",
     "Captures the bottleneck effect at pore-throat junctions — "
     "relevant to whether cells can physically pass between scaffold pores."),

    ("Mean Shortest Path",
     "hops (throats)",
     "Average number of throats traversed on the shortest path between "
     "all pore pairs in the network. Computed via networkx (optional).",
     "Lower values indicate a more directly navigable scaffold — "
     "cells can reach any destination with fewer migration steps.",
     "Captures global network navigability beyond local coordination number."),
]

for ri, row_data in enumerate(metrics_table):
    er = 3 + ri
    bg = C_ALT if ri % 2 == 0 else C_WHITE
    for ci, val in enumerate(row_data, 1):
        c = ws9.cell(row=er, column=ci, value=val)
        c.font      = Font(name=FONT, size=9, bold=(ci == 1))
        c.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True)
        c.fill      = PatternFill("solid", fgColor=bg)
    ws9.row_dimensions[er].height = 60

# =============================================================================
# SAVE
# =============================================================================

excel_path = os.path.join(base_data_dir, excel_filename)
wb.save(excel_path)

log("\n" + "=" * 65)
log("COMPLETE")
log("=" * 65)
log(f"Saved:           {excel_path}")
log(f"Runs processed:  {len(all_summaries)}")
log(f"Skipped:         {len(skipped)}")
log(f"Errors:          {len(errors)}")

if skipped:
    log("\nSkipped:")
    for s, r, reason in skipped:
        log(f"  {s}/{r} — {reason}")
if errors:
    log("\nErrors:")
    for s, r, msg in errors:
        log(f"  {s}/{r} — {msg}")

log(f"\nOpen: {excel_path}")
log("Primary metrics: Sheets 2–7")
log("Archived diameter data: Sheet 8")
log("Import Sheets 4–7 directly into GraphPad Prism as Column Tables")
log("=" * 65)
