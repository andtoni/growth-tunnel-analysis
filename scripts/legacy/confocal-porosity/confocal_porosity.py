# =============================================================================
# Confocal Porosity — CZI Z-Stack Batch Analysis
# =============================================================================
# Author:      Andrea Tonelli (tnland001@myuct.ac.za)
# Institution: University of Cape Town
#
# Description:
#   Reads a folder of Zeiss CZI z-stack files, binarises each stack, and
#   computes porosity (void fraction) using the same porespy call used in
#   the CT pore-network pipeline.
#
#   Results are written to:
#     - porosity_results.csv   — flat CSV (one row per sample, append-safe)
#     - porosity_results.xlsx  — formatted Excel workbook (Prism-ready)
#
#   Voxel size is read directly from CZI XML metadata. If metadata is absent
#   the script falls back to a user-defined default.
#
# Threshold modes:
#   "fixed"  — single global threshold for every file
#   "otsu"   — automatic per-stack Otsu threshold (skimage)
#
# Usage (uv):
#   uv pip install czifile porespy scikit-image openpyxl pandas numpy matplotlib
#   Edit the USER SETTINGS section below, then:
#   python confocal_porosity.py
# =============================================================================

import matplotlib
matplotlib.use('Agg')

import os
import sys
import glob
import xml.etree.ElementTree as ET
import numpy as np
import pandas as pd
import porespy as ps
import czifile
from datetime import datetime
from skimage.filters import threshold_otsu
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# USER SETTINGS — edit only this section
# =============================================================================

# Base folder containing one subfolder per sample
base_folder = r"C:\Users\andto\OneDrive\Desktop\University\PhD\DATA\Transmural Space Characterisation\3D Analysis Paper\Porosity\Confocal"

# Sample names — must match subfolder names exactly
samples = [
    "p16",
    "p18",
    "p20",
]

# Output folder (created automatically)
output_dir = r"C:\Users\andto\OneDrive\Desktop\University\PhD\DATA\Transmural Space Characterisation\3D Analysis Paper\Porosity\Confocal\porosity_output"

# Channel index to analyse (0-based). For single-channel stacks use 0.
# Check your Zeiss acquisition settings if unsure.
channel_index = 0

# Threshold mode: "fixed" or "otsu"
threshold_mode = "otsu"

# Fixed threshold value — only used when threshold_mode = "fixed"
# Pore phase is defined as pixels BELOW this value (same convention as CT scripts)
fixed_threshold = 128

# Voxel size fallback (µm) — used only if CZI metadata does not contain it
# Set this to your objective/zoom pixel size if metadata is missing
fallback_voxel_size_um = 0.21

# Output filenames (saved to output_dir)
csv_filename   = "porosity_results.csv"
excel_filename = "porosity_results.xlsx"

# =============================================================================
# DO NOT EDIT BELOW THIS LINE
# =============================================================================

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
    sys.stdout.flush()


def read_voxel_size_from_czi(czi_path, fallback):
    """
    Parses voxel size (Z, Y, X) in µm from CZI XML metadata.
    Zeiss stores scaling in metres under Metadata/Scaling/Items/Distance.
    Returns (voxel_z, voxel_y, voxel_x) in µm, falling back to `fallback`
    for any axis where metadata is absent or unreadable.
    """
    voxel = {"X": fallback, "Y": fallback, "Z": fallback}
    try:
        with czifile.CziFile(czi_path) as czi:
            meta_xml = czi.metadata()          # returns XML string
        root = ET.fromstring(meta_xml)
        # Path: ImageDocument > Metadata > Scaling > Items > Distance
        for dist in root.iter("Distance"):
            axis = dist.get("Id", "")
            val_el = dist.find("Value")
            if axis in voxel and val_el is not None and val_el.text:
                metres = float(val_el.text)
                voxel[axis] = round(metres * 1e6, 6)   # m → µm
    except Exception as exc:
        log(f"  Metadata parse warning: {exc} — using fallback voxel size")
    return voxel["Z"], voxel["Y"], voxel["X"]


def load_czi_stack(czi_path, channel_idx):
    """
    Loads a CZI file and returns a (Z, Y, X) numpy array for the chosen channel.
    czifile returns an array whose axes are labelled by czi.axes (e.g. 'BCZYX0').
    We locate Z, C, Y, X positions dynamically so the function is robust to
    different acquisition configurations.
    """
    with czifile.CziFile(czi_path) as czi:
        axes  = czi.axes          # e.g. 'TCZYX0' or 'BCZYX0'
        data  = czi.asarray()     # full array, all axes

    # Squeeze size-1 axes except C, Z, Y, X
    axes_list = list(axes)
    data = data.squeeze()

    # After squeeze, rebuild axes string (drop singleton dims)
    squeezed_axes = "".join(
        ax for ax, sz in zip(axes_list, czifile.CziFile(czi_path).shape)
        if sz > 1
    )

    # Fallback: if squeeze removed C (single channel), treat whole array as ZYX
    if "C" not in squeezed_axes:
        # Confirm it looks like ZYX
        if data.ndim == 3:
            return data
        raise ValueError(
            f"Unexpected array shape {data.shape} after squeeze — "
            f"original axes: {axes}"
        )

    # Select channel along C axis
    c_pos = squeezed_axes.index("C")
    n_channels = data.shape[c_pos]
    if channel_idx >= n_channels:
        raise ValueError(
            f"channel_index {channel_idx} out of range "
            f"(file has {n_channels} channel(s))"
        )
    stack = np.take(data, channel_idx, axis=c_pos)   # removes C axis → ZYX
    return stack


# ── Startup log ───────────────────────────────────────────────────────────────
os.makedirs(output_dir, exist_ok=True)

log("=" * 65)
log("Confocal Porosity — CZI Batch Analysis")
log("=" * 65)
log(f"Base folder:      {base_folder}")
log(f"Samples:          {', '.join(samples)}")
log(f"Output dir:       {output_dir}")
log(f"Channel index:    {channel_index}")
log(f"Threshold mode:   {threshold_mode}")
if threshold_mode == "fixed":
    log(f"Fixed threshold:  {fixed_threshold}")
log("=" * 65)

# ── Discover CZI files — one subfolder per sample ─────────────────────────────
# Builds a list of (sample_name, czi_path) pairs.
# Each sample folder is expected to contain one or more .czi files.
# If a folder contains multiple CZI files, all are processed and labelled
# as <sample>_01, <sample>_02, etc.

czi_queue = []   # list of (sample_name, czi_path)

for sample in samples:
    sample_folder = os.path.join(base_folder, sample)

    if not os.path.isdir(sample_folder):
        log(f"WARNING: folder not found for sample '{sample}': {sample_folder}")
        continue

    found = sorted(glob.glob(os.path.join(sample_folder, "*.czi")))

    if not found:
        log(f"WARNING: no .czi files found in {sample_folder}")
        continue

    if len(found) == 1:
        czi_queue.append((sample, found[0]))
    else:
        # Multiple CZI files in the folder — label them _01, _02, …
        for i, path in enumerate(found, 1):
            czi_queue.append((f"{sample}_{i:02d}", path))

if not czi_queue:
    log("ERROR: No CZI files found for any sample — check your paths and sample names")
    sys.exit(1)

log(f"Files queued for analysis ({len(czi_queue)} total):")
for sname, path in czi_queue:
    log(f"  [{sname}]  {os.path.basename(path)}")
log("-" * 65)

# ── Process each file ─────────────────────────────────────────────────────────
results = []
skipped = []
errors  = []

for sample_name, czi_path in czi_queue:
    fname = os.path.basename(czi_path)
    log(f"\nProcessing: [{sample_name}]  {fname}")

    try:
        # ── Load stack ────────────────────────────────────────────────────────
        stack = load_czi_stack(czi_path, channel_index)
        log(f"  Stack shape:     {stack.shape}  (Z, Y, X)")
        log(f"  Stack dtype:     {stack.dtype}")
        log(f"  Intensity range: [{stack.min()}, {stack.max()}]")

        # ── Voxel size from metadata ──────────────────────────────────────────
        voxel_z, voxel_y, voxel_x = read_voxel_size_from_czi(
            czi_path, fallback_voxel_size_um
        )
        log(f"  Voxel size (µm): Z={voxel_z:.4f}, Y={voxel_y:.4f}, X={voxel_x:.4f}")
        voxel_vol_um3 = voxel_z * voxel_y * voxel_x

        # ── Threshold ─────────────────────────────────────────────────────────
        if threshold_mode == "otsu":
            thresh = threshold_otsu(stack)
            log(f"  Otsu threshold:  {thresh:.1f}")
        else:
            thresh = fixed_threshold
            log(f"  Fixed threshold: {thresh}")

        # ── Binarise — pore phase = pixels BELOW threshold ────────────────────
        # Matches CT script convention: im = im3d < threshold
        im_pores = stack < thresh

        # ── Porosity ──────────────────────────────────────────────────────────
        porosity_pct = ps.metrics.porosity(im_pores) * 100
        log(f"  Porosity:        {porosity_pct:.4f}%")

        # ── Volume stats ──────────────────────────────────────────────────────
        n_voxels_total = im_pores.size
        n_voxels_pore  = int(im_pores.sum())
        vol_total_mm3  = (n_voxels_total * voxel_vol_um3) / 1e9
        vol_pore_mm3   = (n_voxels_pore  * voxel_vol_um3) / 1e9
        n_slices, n_rows, n_cols = stack.shape

        results.append({
            "Sample":           sample_name,
            "File":             fname,
            "Z_slices":         n_slices,
            "Y_pixels":         n_rows,
            "X_pixels":         n_cols,
            "Voxel_Z_um":       round(voxel_z, 4),
            "Voxel_Y_um":       round(voxel_y, 4),
            "Voxel_X_um":       round(voxel_x, 4),
            "Threshold_mode":   threshold_mode,
            "Threshold_value":  round(float(thresh), 1),
            "Porosity_pct":     round(porosity_pct, 4),
            "Vol_total_mm3":    round(vol_total_mm3, 6),
            "Vol_pore_mm3":     round(vol_pore_mm3, 6),
            "Processed_at":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    except ValueError as exc:
        log(f"  SKIPPED — {exc}")
        skipped.append((fname, str(exc)))
    except Exception as exc:
        log(f"  ERROR — {exc}")
        errors.append((fname, str(exc)))

log("\n" + "=" * 65)
log(f"Processing complete: {len(results)} succeeded, "
    f"{len(skipped)} skipped, {len(errors)} errors")

if not results:
    log("No results to save — exiting")
    sys.exit(1)

# ── Save CSV ──────────────────────────────────────────────────────────────────
df = pd.DataFrame(results)
csv_path = os.path.join(output_dir, csv_filename)
df.to_csv(csv_path, index=False)
log(f"\nCSV saved:   {csv_path}")

# =============================================================================
# Excel workbook
# =============================================================================

FONT  = "Calibri"
C_HDR = "2E4057"
C_SUB = "048A81"
C_ALT = "EEF7F6"
C_WHT = "FFFFFF"
C_WRN = "FFF3CD"

def hc(ws, row, col, val, bg=C_WHT, fg="000000", sz=10,
       bold=False, italic=False, left=False, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name=FONT, size=sz, bold=bold, italic=italic, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="center", wrap_text=wrap)
    return c

def cw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

wb = Workbook()

# ── Sheet 1: README ───────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "README"
ws1.sheet_view.showGridLines = False
cw(ws1, 1, 40); cw(ws1, 2, 80)

hc(ws1, 1, 1, "Confocal Porosity — README",
   bg=C_HDR, fg=C_WHT, sz=13, bold=True, left=True)
ws1.merge_cells("A1:B1")
ws1.row_dimensions[1].height = 32

readme_rows = [
    ("Generated by",     "confocal_porosity.py"),
    ("Date",             datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ("Base folder",      base_folder),
    ("Samples",          ", ".join(samples)),
    ("Channel analysed", str(channel_index)),
    ("Threshold mode",   threshold_mode),
    ("Pore convention",  "Pixels BELOW threshold = pore (void) phase"),
    ("Porosity formula", "pore voxels / total voxels × 100  [porespy.metrics.porosity]"),
    ("Prism import",     "Copy columns A–B of 'Porosity Summary' as a Column table"),
    ("Files processed",  str(len(results))),
    ("Files skipped",    str(len(skipped))),
    ("Files errored",    str(len(errors))),
]
for ri, (k, v) in enumerate(readme_rows):
    r  = 3 + ri
    bg = C_ALT if ri % 2 == 0 else C_WHT
    hc(ws1, r, 1, k, bg=bg, bold=True, left=True)
    hc(ws1, r, 2, v, bg=bg, left=True)

# ── Sheet 2: Porosity Summary (Prism-ready) ───────────────────────────────────
ws2 = wb.create_sheet("Porosity Summary")
ws2.sheet_view.showGridLines = False
hc(ws2, 1, 1, "Porosity Summary — Confocal Z-Stacks",
   bg=C_HDR, fg=C_WHT, sz=12, bold=True, left=True)
ws2.merge_cells("A1:D1")
ws2.row_dimensions[1].height = 28

for ci, (h, w) in enumerate(zip(
        ["Sample", "Porosity (%)", "Threshold", "Threshold Mode"],
        [32, 18, 16, 18]), 1):
    hc(ws2, 2, ci, h, bg=C_SUB, fg=C_WHT, bold=True)
    cw(ws2, ci, w)

for ri, row in enumerate(results):
    r  = 3 + ri
    bg = C_ALT if ri % 2 == 0 else C_WHT
    hc(ws2, r, 1, row["Sample"],          bg=bg, left=True)
    hc(ws2, r, 2, row["Porosity_pct"],    bg=bg)
    hc(ws2, r, 3, row["Threshold_value"], bg=bg)
    hc(ws2, r, 4, row["Threshold_mode"],  bg=bg)

# ── Sheet 3: Full Data ────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Full Data")
ws3.sheet_view.showGridLines = False
hc(ws3, 1, 1, "Full Data — All Metrics",
   bg=C_HDR, fg=C_WHT, sz=12, bold=True, left=True)
ws3.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
ws3.row_dimensions[1].height = 28

col_widths_full = [28, 32, 10, 10, 10, 12, 12, 12, 16, 16, 14, 14, 14, 22]
for ci, (col, w) in enumerate(zip(df.columns, col_widths_full), 1):
    hc(ws3, 2, ci, col, bg=C_SUB, fg=C_WHT, bold=True, sz=9)
    cw(ws3, ci, w)
for ri, row in enumerate(results):
    r  = 3 + ri
    bg = C_ALT if ri % 2 == 0 else C_WHT
    for ci, col in enumerate(df.columns, 1):
        hc(ws3, r, ci, row[col], bg=bg, left=(ci == 1), sz=9)

# ── Sheet 4: Errors & Skipped ─────────────────────────────────────────────────
if skipped or errors:
    ws4 = wb.create_sheet("Errors and Skipped")
    ws4.sheet_view.showGridLines = False
    hc(ws4, 1, 1, "Errors and Skipped Files",
       bg=C_HDR, fg=C_WHT, sz=12, bold=True, left=True)
    ws4.merge_cells("A1:C1")
    cw(ws4, 1, 36); cw(ws4, 2, 14); cw(ws4, 3, 60)
    r = 2
    if skipped:
        hc(ws4, r, 1, "SKIPPED", bg=C_WRN, bold=True, left=True); r += 1
        for fn, reason in skipped:
            hc(ws4, r, 1, fn,     bg=C_WRN, left=True)
            hc(ws4, r, 3, reason, bg=C_WRN, left=True, wrap=True); r += 1
    if errors:
        hc(ws4, r, 1, "ERRORS", bg="FFCCCC", bold=True, left=True); r += 1
        for fn, msg in errors:
            hc(ws4, r, 1, fn,  bg="FFCCCC", left=True)
            hc(ws4, r, 3, msg, bg="FFCCCC", left=True, wrap=True); r += 1

# ── Save ──────────────────────────────────────────────────────────────────────
excel_path = os.path.join(output_dir, excel_filename)
wb.save(excel_path)
log(f"Excel saved: {excel_path}")

# ── Final summary ─────────────────────────────────────────────────────────────
log("\n" + "=" * 65)
log("RESULTS SUMMARY")
log("=" * 65)
log(f"{'Sample':<35} {'Porosity (%)':>14}  {'Threshold':>12}")
log("-" * 65)
for row in results:
    log(f"{row['Sample']:<35} {row['Porosity_pct']:>14.4f}  "
        f"{row['Threshold_value']:>12.1f}")

if skipped:
    log("\nSkipped:")
    for fn, reason in skipped:
        log(f"  {fn} — {reason}")
if errors:
    log("\nErrors:")
    for fn, msg in errors:
        log(f"  {fn} — {msg}")

log("\n" + "=" * 65)
log("Output files:")
log(f"  {csv_path}")
log(f"  {excel_path}")
log("=" * 65)
