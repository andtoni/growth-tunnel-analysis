# =============================================================================
# PNM Pore & Throat Data Organiser for GraphPad Prism
# =============================================================================
# Author:      Andrea Tonelli (tnland001@myuct.ac.za)
# ORCID:       https://orcid.org/0000-0002-1601-4103
# Institution: University of Cape Town
# Repository:  https://github.com/andtoni/growth-tunnel-analysis
#
# Description:
#   Reads all pore and throat Excel files from INPUT_DIR, takes the first
#   N_ROWS measurements of each numeric variable, and writes a single
#   output Excel workbook formatted for import into GraphPad Prism as a
#   Grouped table.
#
#   File naming convention expected:
#     pores<sample>.xlsx   e.g. pores16SR.xlsx, pores20HR.xlsx
#     throats<sample>.xlsx e.g. throats16SR.xlsx, throats20HR.xlsx
#
#   Output sheet per variable:
#     Each column  = one sample (e.g. 16SR, 16HR)
#     Each row     = one measurement (up to N_ROWS)
#   Ready to import into Prism: File → Import → From File → Grouped table
#
# Usage:
#   python pnm_prism_export.py
# =============================================================================

import os
import re
import glob
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# =============================================================================
# USER SETTINGS
# =============================================================================

# Directory containing all pore*.xlsx and throat*.xlsx files
INPUT_DIR = r"C:\Users\andto\OneDrive\Desktop\University\PhD\DATA\Transmural Space Characterisation\3D Analysis Paper\PNM\Pore_Throats_Summary"

# Number of measurements to take from the top of each file
N_ROWS = 512

# Output filename — saved into INPUT_DIR
OUTPUT_FILENAME = "PNM_Prism_Grouped.xlsx"

# Columns to exclude (metadata, not measurements)
EXCLUDE_COLS = {"group", "weight", "sample", "id", "index"}

# =============================================================================
# DO NOT EDIT BELOW THIS LINE
# =============================================================================

def parse_filename(fname):
    """
    Extract data type and sample name from filename.
    e.g. pores16SR.xlsx   → ('pore',   '16SR')
         throats20HR.xlsx → ('throat', '20HR')
    Returns (data_type, sample_name) or (None, None) if not matched.
    """
    base = os.path.splitext(os.path.basename(fname))[0]
    m = re.match(r'^pores?(.+)$', base, re.IGNORECASE)
    if m:
        return 'pore', m.group(1)
    m = re.match(r'^throats?(.+)$', base, re.IGNORECASE)
    if m:
        return 'throat', m.group(1)
    return None, None


def load_file(fpath, n_rows):
    """
    Load an Excel file, convert all data columns to numeric,
    and return the first n_rows rows of data columns only.
    """
    df = pd.read_excel(fpath)

    # Drop metadata/excluded columns
    data_cols = [c for c in df.columns
                 if c.lower() not in EXCLUDE_COLS]

    df = df[data_cols].copy()

    # Coerce all columns to numeric (drops non-numeric entries silently)
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    # Drop rows where ALL values are NaN
    df.dropna(how='all', inplace=True)
    df.reset_index(drop=True, inplace=True)

    return df.head(n_rows)


# =============================================================================
# STYLE HELPERS
# =============================================================================

FONT      = "Arial"
C_PORE    = "1F6AA5"   # deep blue  — pore sheets
C_THROAT  = "B85C00"   # burnt orange — throat sheets
C_ALT     = "D6E4F0"   # light blue — alternating rows
C_WHITE   = "FFFFFF"
C_HDR_FG  = "FFFFFF"
C_DARK    = "1A1A1A"


def hdr_cell(ws, row, col, value, bg, fg=C_HDR_FG, sz=10, bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=FONT, bold=bold, color=fg, size=sz)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    return c


def dat_cell(ws, row, col, value, bg=C_WHITE):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=FONT, size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.number_format = "0.0000"
    return c


def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


# =============================================================================
# MAIN
# =============================================================================

print("=" * 60)
print("PNM Prism Export")
print("=" * 60)
print(f"Input directory : {INPUT_DIR}")
print(f"Rows per sample : {N_ROWS}")

# ── Discover all matching files ───────────────────────────────────────────────
all_files = sorted(glob.glob(os.path.join(INPUT_DIR, "*.xlsx")))
# Exclude any previously generated output file
all_files = [f for f in all_files
             if os.path.basename(f) != OUTPUT_FILENAME]

pore_files   = {}   # sample_name → filepath
throat_files = {}

for fpath in all_files:
    dtype, sample = parse_filename(fpath)
    if dtype == 'pore':
        pore_files[sample] = fpath
        print(f"  Pore   [{sample}]: {os.path.basename(fpath)}")
    elif dtype == 'throat':
        throat_files[sample] = fpath
        print(f"  Throat [{sample}]: {os.path.basename(fpath)}")
    else:
        print(f"  Skipped (unrecognised name): {os.path.basename(fpath)}")

if not pore_files and not throat_files:
    raise FileNotFoundError(
        f"No pore or throat files found in:\n  {INPUT_DIR}\n"
        "Files must be named pores<sample>.xlsx or throats<sample>.xlsx"
    )

# ── Load data ─────────────────────────────────────────────────────────────────
pore_data   = {}   # sample → DataFrame
throat_data = {}   # sample → DataFrame

for sample, fpath in sorted(pore_files.items()):
    pore_data[sample] = load_file(fpath, N_ROWS)
    print(f"  Loaded pore [{sample}]: "
          f"{len(pore_data[sample])} rows, "
          f"cols: {pore_data[sample].columns.tolist()}")

for sample, fpath in sorted(throat_files.items()):
    throat_data[sample] = load_file(fpath, N_ROWS)
    print(f"  Loaded throat [{sample}]: "
          f"{len(throat_data[sample])} rows, "
          f"cols: {throat_data[sample].columns.tolist()}")

# Collect all unique variable names per type
pore_vars   = []
throat_vars = []
for df in pore_data.values():
    for c in df.columns:
        if c not in pore_vars:
            pore_vars.append(c)
for df in throat_data.values():
    for c in df.columns:
        if c not in throat_vars:
            throat_vars.append(c)

pore_samples   = sorted(pore_data.keys())
throat_samples = sorted(throat_data.keys())

print(f"\nPore variables   : {pore_vars}")
print(f"Throat variables : {throat_vars}")
print(f"Pore samples     : {pore_samples}")
print(f"Throat samples   : {throat_samples}")

# =============================================================================
# BUILD EXCEL WORKBOOK
# =============================================================================

wb = Workbook()
wb.remove(wb.active)   # remove default blank sheet

# ── README sheet ──────────────────────────────────────────────────────────────
ws_readme = wb.create_sheet("README")
ws_readme.sheet_view.showGridLines = False

readme_rows = [
    ("PNM Pore & Throat Data — GraphPad Prism Import",),
    ("",),
    ("Author",        "Andrea Tonelli — University of Cape Town"),
    ("ORCID",         "https://orcid.org/0000-0002-1601-4103"),
    ("Repository",    "https://github.com/andtoni/growth-tunnel-analysis"),
    ("Generated",     datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("Input folder",  INPUT_DIR),
    ("Rows per sample", str(N_ROWS)),
    ("",),
    ("HOW TO IMPORT INTO GRAPHPAD PRISM",),
    ("Step 1", "Open GraphPad Prism → New → Grouped Table"),
    ("Step 2", "File → Import → From File → select this workbook"),
    ("Step 3", "Choose the sheet for the variable you want to analyse"),
    ("Step 4", "Each column is one sample. Each row is one measurement."),
    ("Step 5", "Use Analyze → Column Statistics for descriptive stats,"),
    ("",        "or Analyze → One-way ANOVA for group comparisons."),
    ("",),
    ("SHEET GUIDE",),
    ("Pore sheets (blue)",   "One sheet per pore measurement variable"),
    ("Throat sheets (orange)", "One sheet per throat measurement variable"),
    ("",),
    ("NOTE",
     f"Only the first {N_ROWS} measurements per sample are included. "
     "Samples with fewer rows will have blank cells at the bottom."),
]

ws_readme.column_dimensions['A'].width = 20
ws_readme.column_dimensions['B'].width = 75

for ri, row_data in enumerate(readme_rows, 1):
    is_title    = len(row_data) == 1 and row_data[0] != ""
    is_section  = row_data[0] in ("HOW TO IMPORT INTO GRAPHPAD PRISM",
                                   "SHEET GUIDE")
    for ci, val in enumerate(row_data, 1):
        c = ws_readme.cell(row=ri, column=ci, value=val)
        if ri == 1:
            c.font = Font(name=FONT, bold=True, size=13,
                          color=C_PORE)
        elif is_section:
            c.font = Font(name=FONT, bold=True, size=10,
                          color=C_HDR_FG)
            c.fill = PatternFill("solid", fgColor=C_PORE)
        elif is_title:
            c.font = Font(name=FONT, bold=True, size=10)
        else:
            c.font = Font(name=FONT, size=9,
                          bold=(ci == 1 and row_data[0] != ""))
        c.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True)
    ws_readme.row_dimensions[ri].height = 18 if not is_title else 20


def build_variable_sheet(wb, sheet_name, variable, samples,
                          data_dict, colour):
    """
    Creates one Prism-ready grouped sheet for a single variable.
    Columns = samples, rows = measurements (up to N_ROWS).
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    n_samples = len(samples)

    # ── Title row ─────────────────────────────────────────────────────────────
    title_val = f"{variable}   |   {sheet_name}   |   n = {N_ROWS} per sample"
    hdr_cell(ws, 1, 1, title_val, bg=colour, sz=11)
    if n_samples > 1:
        ws.merge_cells(
            f"A1:{get_column_letter(n_samples)}1"
        )
    ws.row_dimensions[1].height = 24

    # ── Import hint row ────────────────────────────────────────────────────────
    hint = ("Import into Prism as Grouped Table — "
            "each column = one sample, each row = one measurement")
    c = ws.cell(row=2, column=1, value=hint)
    c.font      = Font(name=FONT, italic=True, size=8, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")
    if n_samples > 1:
        ws.merge_cells(f"A2:{get_column_letter(n_samples)}2")
    ws.row_dimensions[2].height = 14

    # ── Column headers (sample names) ─────────────────────────────────────────
    for ci, sample in enumerate(samples, 1):
        hdr_cell(ws, 3, ci, sample, bg=colour, sz=10)
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.row_dimensions[3].height = 30

    # ── Unit sub-header ────────────────────────────────────────────────────────
    for ci in range(1, n_samples + 1):
        c = ws.cell(row=4, column=ci, value="µm" if "diameter" in variable
                    else ("µm³" if variable == "volume" else ""))
        c.font      = Font(name=FONT, italic=True, size=8, color="595959")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = PatternFill("solid", fgColor="F2F2F2")
    ws.row_dimensions[4].height = 13

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx in range(N_ROWS):
        excel_row = 5 + row_idx
        bg = C_ALT if row_idx % 2 == 0 else C_WHITE
        for ci, sample in enumerate(samples, 1):
            df = data_dict.get(sample)
            if df is not None and variable in df.columns and row_idx < len(df):
                val = df[variable].iloc[row_idx]
                val = float(val) if pd.notna(val) else None
            else:
                val = None
            dat_cell(ws, excel_row, ci, val, bg=bg)

    # Freeze header rows
    ws.freeze_panes = f"A5"
    return ws


# ── Pore sheets ───────────────────────────────────────────────────────────────
for var in pore_vars:
    sheet_name = f"Pore — {var}"
    print(f"  Writing sheet: {sheet_name}")
    build_variable_sheet(wb, sheet_name, var,
                         pore_samples, pore_data, C_PORE)

# ── Throat sheets ─────────────────────────────────────────────────────────────
for var in throat_vars:
    sheet_name = f"Throat — {var}"
    print(f"  Writing sheet: {sheet_name}")
    build_variable_sheet(wb, sheet_name, var,
                         throat_samples, throat_data, C_THROAT)

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = os.path.join(INPUT_DIR, OUTPUT_FILENAME)
wb.save(out_path)

print()
print("=" * 60)
print("COMPLETE")
print(f"  Output: {out_path}")
print(f"  Pore variables   : {pore_vars}")
print(f"  Throat variables : {throat_vars}")
print(f"  Pore samples     : {pore_samples}")
print(f"  Throat samples   : {throat_samples}")
print("=" * 60)
